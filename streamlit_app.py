import os
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

data_atual = datetime.now().strftime("%Y-%m-%d")


def carregar_planilha(file, skiprows):
    try:
        planilha = pd.read_excel(file, skiprows=skiprows)
        return planilha
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo: {e}")
        return None

def estilizar_dataframe(df, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Adicionar cabe√ßalho
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Estilo do cabe√ßalho
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border

    # Estilo das c√©lulas
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    return wb

def to_excel_bytes(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
    
def processar_planilha_simplificada(file):
    try:
        df = pd.read_excel(file, header=7)
        df = df[['Medicamento', 'Lote', 'Data Vencimento', 'Quantidade Encontrada']]
        return df.dropna()
    except Exception as e:
        st.warning(f"Erro ao processar {file.name}: {e}")
        return None

def processar_e_juntar_planilhas(pasta_raw):
    lista_dfs = []

    for nome_arquivo in os.listdir(pasta_raw):
        if nome_arquivo.endswith(('.xlsx', '.xls')):
            caminho_arquivo = os.path.join(pasta_raw, nome_arquivo)
            try:
                df = pd.read_excel(caminho_arquivo, header=12)
                df = df[['CodAuxiliar - Produto / Fabricante / Marca / Embalagem',
                         'Lote', 'Validade', 'Endere√ßo', 'Posi√ß√£o', 'Cont. 1']]
                df['Nome Medicamento'] = df['CodAuxiliar - Produto / Fabricante / Marca / Embalagem'] \
                    .str.extract(r'-\s*(.*?)\s*\[')
                df = df[['Endere√ßo', 'Posi√ß√£o', 'Nome Medicamento', 'Lote', 'Validade', 'Cont. 1']]
                df['Planilha'] = nome_arquivo
                df = df.dropna()
                lista_dfs.append(df)
            except Exception as e:
                st.warning(f"Erro ao processar {nome_arquivo}: {e}")

    if lista_dfs:
        df_final = pd.concat(lista_dfs, ignore_index=True)
        return df_final
    else:
        return None


def filtrar_maior_id_por_posicao(df):
    colunas = ['IDListaInventario', 'NMEndereco', 'CDPosicao', 'NMProduto', 'CDLote', 'QTFinal']
    df = df[colunas]
    df = df.sort_values('IDListaInventario', ascending=False)
    return df.drop_duplicates(subset='CDPosicao', keep='first')


def comparacao_hosp(df_hosp, df_sesab):
    df_hosp = df_hosp.sort_values('IDListaInventario', ascending=False)
    df_filtrado = df_hosp.drop_duplicates(subset='CDPosicao', keep='first')

    df_filtrado = df_filtrado.rename(columns={
        'NMEndereco': 'Endere√ßo',
        'CDPosicao': 'Posi√ß√£o',
        'CDLote': 'Lote',
        'QTFinal': 'Contagem Hosplog'
    })

    df_filtrado = df_filtrado[['Posi√ß√£o', 'Endere√ßo', 'Lote', 'Contagem Hosplog']]
    return pd.merge(df_sesab, df_filtrado, how='outer', on=['Posi√ß√£o', 'Lote'])


# --- Streamlit App ---
st.title("üìä Processador de Planilhas de Invent√°rio")

with st.expander("1. Processar planilhas EspelhoInventario Hosplog"):
    st.markdown("### üìÇ Selecione as planilhas EspelhoInventario (.xlsx/.xls)")

    arquivos_selecionados = st.file_uploader(
        "Selecione uma ou mais planilhas",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="espelho_inventario"
    )

    if arquivos_selecionados:
        lista_dfs = []

        for arquivo in arquivos_selecionados:
            try:
                df = pd.read_excel(arquivo, header=12)
                df = df[['CodAuxiliar - Produto / Fabricante / Marca / Embalagem',
                         'Lote', 'Validade', 'Endere√ßo', 'Posi√ß√£o', 'Cont. 1']]
                
                df['Nome Medicamento'] = df['CodAuxiliar - Produto / Fabricante / Marca / Embalagem'] \
                    .str.extract(r'-\s*(.*?)\s*\[')
                
                df = df[['Endere√ßo', 'Posi√ß√£o', 'Nome Medicamento', 'Lote', 'Validade', 'Cont. 1']]
                df['Planilha'] = arquivo.name
                df = df.dropna()
                
                lista_dfs.append(df)
            except Exception as e:
                st.warning(f"Erro ao processar {arquivo.name}: {e}")

        if lista_dfs:
            df_unificado = pd.concat(lista_dfs, ignore_index=True)
            st.success(f"‚úÖ {len(df_unificado)} registros processados com sucesso!")
            st.dataframe(df_unificado.head())

            buffer = BytesIO()
            df_unificado.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)

            st.download_button("üì• Baixar Planilha Unificada", buffer,
                               file_name="planilha_unificada.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("‚ö†Ô∏è Nenhum dado v√°lido foi processado.")


with st.expander("2. Filtro por √öltimo ID (Hosplog) - Planilha inventario"):
    planilha_hosp = st.file_uploader("Carregue a planilha da Hosplog", type=["xlsx"])
    if planilha_hosp:
            df_hosp = pd.read_excel(planilha_hosp)
            df_filtrado = filtrar_maior_id_por_posicao(df_hosp)
            st.success("Filtro aplicado com sucesso!")
            st.dataframe(df_filtrado.head())
        
            buffer = io.BytesIO()
            df_filtrado.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            st.download_button("üì• Baixar Filtro Hosplog", buffer,
                               file_name="filtrado_hosplog.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


with st.expander("3. Compara√ß√£o Hosplog x Sesab"):
    col1, col2 = st.columns(2)
    with col1:
        planilha_hosp = st.file_uploader("Hosplog", type=["xlsx"], key="hosplog_cmp")
    with col2:
        planilha_sesab = st.file_uploader("Sesab", type=["xlsx"], key="sesab_cmp")

    if planilha_hosp and planilha_sesab:
        df_hosp = pd.read_excel(planilha_hosp)
        df_sesab = pd.read_excel(planilha_sesab)
        df_cruzado = comparacao_hosp(df_hosp, df_sesab)
        
        st.success("Compara√ß√£o realizada com sucesso!")
        st.dataframe(df_cruzado.head())
        
        buffer = io.BytesIO()
        df_cruzado.to_excel(buffer, index=False, engine='openpyxl')
        buffer.seek(0)
        st.download_button("üì• Baixar Cruzamento", buffer,
                       file_name="cruzamento_hosp_sesab.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
with st.expander("4. Processar Planilha de Estoque AFSESAB (Cabe√ßalho come√ßando na 8 linha)"):
    arquivo_simples = st.file_uploader("Selecione um arquivo .xls (Simples)", type=["xls"], key="planilha_simples")

    if arquivo_simples:
        df_simples = processar_planilha_simplificada(arquivo_simples)
        if df_simples is not None:
            st.success("Planilha processada com sucesso!")
            st.dataframe(df_simples.head())
        
            buffer = BytesIO()
            df_simples.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            st.download_button("üì• Baixar Planilha Processada", buffer,
                               file_name="planilha_simples.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with st.expander("5. Gerar apura√ß√£o SIMPAS"):
    st.subheader("Gerar Apura√ß√£o SIMPAS")
    item_selecionado3 = st.text_input("Nome da Lista:")
    estoque_file3 = st.file_uploader(
        "Upload da planilha de Estoque Final:", type=["xls"]
    )
    if estoque_file3:

        estoque_df = carregar_planilha(estoque_file3, skiprows=7)
        estoque_df = estoque_df[
            [
                "C√≥digo Simpas",
                "Medicamento",
                "Quantidade Encontrada",
                "Programa Sa√∫de",
            ]
        ]
        estoque_df["C√≥digo Simpas"] = estoque_df["C√≥digo Simpas"].astype(str)
        df = (
            estoque_df.groupby(
                [
                    "C√≥digo Simpas",
                    "Medicamento",
                    "Programa Sa√∫de",
                ]
            )["Quantidade Encontrada"]
            .sum()
            .reset_index()
        )
        df = df.sort_values(by="C√≥digo Simpas")
        df = df.rename(
            columns={
                "Quantidade Encontrada": "Quantidade",
            }
        )
        new = ["C√≥digo Simpas", "Medicamento", "Quantidade", "Programa Sa√∫de"]
        df = df[new]

        # Estilizar o DataFrame
        wb = estilizar_dataframe(df, "Apura√ß√£o SIMPAS")
        excel_bytes = to_excel_bytes(wb)

        # Exibir tabelas resultantes
        st.write("Resultado da An√°lise:")
        st.dataframe(df)

        # Bot√£o de download
        st.download_button(
            label="Baixar Planilha de Apura√ß√£o SIMPAS",
            data=excel_bytes,
            file_name=f"{item_selecionado3} Apuracao_SIMPAS {data_atual}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )






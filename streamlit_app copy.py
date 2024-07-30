import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile


def carregar_planilha(file, skiprows=0):
    """Carrega uma planilha Excel"""
    try:
        return pd.read_excel(file, skiprows=skiprows)
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo: {e}")
        return None


def carregar_todas_abas(file):
    """Carrega todas as abas de uma planilha Excel"""
    try:
        xls = pd.ExcelFile(file)
        return pd.concat(
            pd.read_excel(xls, sheet_name=sheet) for sheet in xls.sheet_names
        )
    except Exception as e:
        st.error(f"Erro ao carregar o arquivo com múltiplas abas: {e}")
        return None


def estilizar_dataframe(df, sheet_name):
    """Estiliza um DataFrame e retorna um Workbook"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws["1:1"]:
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border

    return wb


def to_excel_bytes(wb):
    """Converte um Workbook para bytes"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def criar_arquivo_zip(arquivos):
    """Cria um arquivo ZIP a partir de uma lista de arquivos"""
    buffer_zip = BytesIO()
    with zipfile.ZipFile(buffer_zip, "w") as zipf:
        for nome_arquivo, dados in arquivos:
            zipf.writestr(nome_arquivo, dados.getvalue())
    buffer_zip.seek(0)
    return buffer_zip


def gerar_planilha_conferencia(df, nome):
    """Gera a planilha de conferência a partir de um DataFrame"""
    df_conferencia = df.copy()
    for i in range(1, 4):
        df_conferencia[f"Contagem {i}"] = df_conferencia["Contagem"]
    df_conferencia["Valor Adotado"] = df_conferencia["Contagem"]
    df_conferencia = df_conferencia[
        [
            "Medicamento",
            "Lote",
            "Data Vencimento",
            "Contagem 1",
            "Contagem 2",
            "Contagem 3",
            "Valor Adotado",
        ]
    ]
    return estilizar_dataframe(df_conferencia, nome)


def main():
    st.title("Sistema de Inventário")

    opcao = st.selectbox(
        "Escolha uma opção:",
        ["Gerar lista de contagem", "Gerar apuração SIGAF", "Gerar apuração SIMPAS"],
    )

    if opcao == "Gerar lista de contagem":
        st.subheader("Gerar Lista de Contagem")

        item_selecionado = st.text_input("Nome da Lista:")
        estoque_file1 = st.file_uploader(
            "Upload da planilha de Estoque:", type=["xlsx"]
        )
        enderecos_file = st.file_uploader(
            "Upload da planilha de Endereços:", type=["xlsx"]
        )

        if estoque_file1 and enderecos_file:
            estoque_df = carregar_planilha(estoque_file1, skiprows=7)
            enderecos_df = carregar_todas_abas(enderecos_file)

            if estoque_df is not None and enderecos_df is not None:
                data_atual = datetime.now().strftime("%Y%m%d")
                nome_arquivo_estoque = f"Estoque_{item_selecionado}_{data_atual}.xlsx"

                if "Contagem" not in estoque_df.columns:
                    estoque_df["Contagem"] = None

                estoque_df = estoque_df[
                    ["Medicamento", "Lote", "Data Vencimento", "Contagem"]
                ]
                estoque_df["Lote"] = estoque_df["Lote"].astype(str)

                enderecos_df = enderecos_df.rename(
                    columns={
                        "LOCALIZAÇÃO": "Endereço",
                        "PROGRAMA": "Programa",
                        "LOTE": "Lote",
                    }
                )
                enderecos_df["Endereço"].fillna(enderecos_df["05/26"], inplace=True)
                enderecos_df["Lote"] = enderecos_df["Lote"].astype(str).str.rstrip()
                enderecos_df = enderecos_df[["Endereço", "Programa", "Lote"]]

                merged_df = pd.merge(estoque_df, enderecos_df, on="Lote", how="left")
                colunas_reordenadas = [
                    "Endereço",
                    "Medicamento",
                    "Lote",
                    "Data Vencimento",
                    "Programa",
                    "Contagem",
                ]
                merged_df = merged_df[colunas_reordenadas].sort_values(by="Medicamento")

                nome_arquivo_1 = f"{item_selecionado}_contagem_{data_atual}.xlsx"
                nome_arquivo_conferencia = (
                    f"{item_selecionado}_{data_atual}_conferencia.xlsx"
                )
                nome_arquivo_enderecaemento = (
                    f"{item_selecionado}_{data_atual}_endereco.xlsx"
                )

                wb1 = estilizar_dataframe(merged_df, "Contagem")
                wb_conferencia = gerar_planilha_conferencia(merged_df, "Conferência")
                wb_endereco = estilizar_dataframe(enderecos_df, "Endereço")
                wb_estoque = estilizar_dataframe(estoque_df, "Estoque")

                excel_bytes_estoque = to_excel_bytes(wb_estoque)
                excel_bytes = to_excel_bytes(wb1)
                excel_bytes_conferencia = to_excel_bytes(wb_conferencia)
                excel_bytes_enderecaemento = to_excel_bytes(wb_endereco)

                arquivos = [
                    (nome_arquivo_estoque, excel_bytes_estoque),
                    (nome_arquivo_1, excel_bytes),
                    (nome_arquivo_conferencia, excel_bytes_conferencia),
                    (nome_arquivo_enderecaemento, excel_bytes_enderecaemento),
                ]
                arquivo_zip_bytes = criar_arquivo_zip(arquivos)

                st.write("Resultado da Análise:")
                st.dataframe(merged_df)

                st.download_button(
                    label="Baixar Todos os Arquivos",
                    data=arquivo_zip_bytes,
                    file_name=f"{item_selecionado}_{data_atual}_arquivos.zip",
                    mime="application/zip",
                )

    elif opcao == "Gerar apuração SIGAF":
        st.subheader("Gerar Apuração SIGAF")
        estoque_file2 = st.file_uploader(
            "Upload da planilha de Estoque(Gerada):", type=["xlsx"]
        )
        conferencia_file = st.file_uploader(
            "Upload da planilha de Conferencia:", type=["xlsx"]
        )

        if estoque_file2 and conferencia_file:
            conferencia_df = carregar_planilha(conferencia_file)
            conferencia_df = (
                conferencia_df.groupby(["Medicamento", "Lote", "Data Vencimento"])[
                    "Valor Adotado"
                ]
                .sum()
                .reset_index()
            )
            estoque_df = carregar_planilha(estoque_file2)
            estoque_df = (
                estoque_df.groupby(
                    [
                        "Código Simpas",
                        "Medicamento",
                        "Lote",
                        "Data Vencimento",
                        "Valor Unitário",
                        "Programa Saúde",
                    ]
                )["Quantidade Encontrada"]
                .sum()
                .reset_index()
            )

            df = pd.merge(
                conferencia_df,
                estoque_df,
                on=["Lote", "Medicamento", "Data Vencimento"],
                how="left",
            )
            df = df.rename(
                columns={
                    "Data Vencimento": "Validade",
                    "Quantidade Encontrada": "SIGAF",
                    "Valor Adotado": "Contagem",
                }
            )
            df["Diferença"] = df["Contagem"] - df["SIGAF"]
            df["Vlr Total"] = df["Contagem"] * df["Valor Unitário"]
            df["Vlr Divergencia"] = df["Diferença"] * df["Valor Unitário"]

            df = df[
                [
                    "Código Simpas",
                    "Medicamento",
                    "Lote",
                    "Validade",
                    "Contagem",
                    "SIGAF",
                    "Diferença",
                    "Valor Unitário",
                    "Vlr Total",
                    "Vlr Divergencia",
                    "Programa Saúde",
                ]
            ]

            wb = estilizar_dataframe(df, "Apuração")
            excel_bytes = to_excel_bytes(wb)

            st.write("Resultado da Análise:")
            st.dataframe(df)

            st.download_button(
                label="Baixar Planilha de Apuração",
                data=excel_bytes,
                file_name=f"Apuracao_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    elif opcao == "Gerar apuração SIMPAS":
        st.subheader("Gerar Apuração SIMPAS")
        estoque_file3 = st.file_uploader(
            "Upload da planilha de Estoque Final:", type=["xlsx"]
        )

        if estoque_file3:
            estoque_df = carregar_planilha(estoque_file3, skiprows=7)
            estoque_df = (
                estoque_df.groupby(["Código Simpas", "Medicamento", "Programa Saúde"])[
                    "Quantidade Encontrada"
                ]
                .sum()
                .reset_index()
            )
            estoque_df = estoque_df.sort_values(by="Código Simpas")
            estoque_df = estoque_df.rename(
                columns={"Quantidade Encontrada": "Quantidade"}
            )

            wb = estilizar_dataframe(estoque_df, "Apuração SIMPAS")
            excel_bytes = to_excel_bytes(wb)

            st.write("Resultado da Análise:")
            st.dataframe(estoque_df)

            st.download_button(
                label="Baixar Planilha de Apuração SIMPAS",
                data=excel_bytes,
                file_name=f"Apuracao_SIMPAS_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
